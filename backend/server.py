from fastapi import FastAPI, APIRouter, HTTPException, Depends, status
from contextlib import asynccontextmanager
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse, RedirectResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
import json
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Configuration
JWT_SECRET = os.environ.get('JWT_SECRET', 'your-secret-key-change-in-production')
JWT_ALGORITHM = 'HS256'
JWT_EXPIRATION_HOURS = 24

# Lifespan context manager
@asynccontextmanager
async def lifespan(app: FastAPI):
    # Startup
    yield
    # Shutdown
    client.close()

# Create the main app without a prefix
app = FastAPI(lifespan=lifespan)

@app.get("/", include_in_schema=False)
async def root_redirect():
    return RedirectResponse(url="/docs")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

security = HTTPBearer()

# ============ MODELS ============

class TutorRegister(BaseModel):
    email: EmailStr
    password: str
    name: str

class TutorLogin(BaseModel):
    email: EmailStr
    password: str

class Tutor(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: str
    name: str
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class QuestionCreate(BaseModel):
    type: str  # 'multiple_choice', 'true_false', 'short_answer'
    question_text: str
    options: Optional[List[str]] = None
    correct_answer: str
    points: int = 1
    randomize_options: bool = True

class Question(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    type: str
    question_text: str
    options: Optional[List[str]] = None
    correct_answer: str
    points: int
    randomize_options: bool

class ExamSettings(BaseModel):
    require_fullscreen: bool = True
    detect_tab_switch: bool = True
    disable_copy_paste: bool = True
    disable_right_click: bool = True
    max_violations: int = 3
    randomize_questions: bool = True
    show_results_immediately: bool = False
    time_limit_minutes: Optional[int] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None

class ExamCreate(BaseModel):
    title: str
    description: str
    required_fields: List[str]  # ['name', 'email', 'student_id', etc.]
    questions: List[QuestionCreate]
    settings: ExamSettings

class Exam(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tutor_id: str
    title: str
    description: str
    required_fields: List[str]
    questions: List[Question]
    settings: ExamSettings
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    is_active: bool = True

class ViolationLog(BaseModel):
    type: str  # 'tab_switch', 'fullscreen_exit', 'copy_attempt', 'right_click', etc.
    timestamp: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    details: Optional[str] = None

class StudentAnswer(BaseModel):
    question_id: str
    answer: str
    time_spent_seconds: int

class ExamSubmission(BaseModel):
    exam_id: str
    student_data: Dict[str, str]  # Dynamic fields like name, email, student_id
    answers: List[StudentAnswer]
    violations: List[ViolationLog]
    browser_info: Optional[Dict[str, Any]] = None
    ip_address: Optional[str] = None

class ExamAttempt(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    exam_id: str
    student_data: Dict[str, str]
    answers: List[StudentAnswer]
    violations: List[ViolationLog]
    score: float
    max_score: int
    percentage: float
    submitted_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    flagged: bool = False
    browser_info: Optional[Dict[str, Any]] = None
    ip_address: Optional[str] = None

class ViolationReport(BaseModel):
    exam_id: str
    violation: ViolationLog

# ============ AUTH HELPERS ============

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_jwt_token(tutor_id: str) -> str:
    payload = {
        'tutor_id': tutor_id,
        'exp': datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_tutor(credentials: HTTPAuthorizationCredentials = Depends(security)) -> str:
    try:
        token = credentials.credentials
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        tutor_id = payload.get('tutor_id')
        if not tutor_id:
            raise HTTPException(status_code=401, detail="Invalid token")
        return tutor_id
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# ============ TUTOR ROUTES ============

@api_router.post("/tutors/register")
async def register_tutor(tutor_data: TutorRegister):
    # Check if tutor exists
    existing = await db.tutors.find_one({"email": tutor_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Create tutor
    tutor_obj = Tutor(
        email=tutor_data.email,
        name=tutor_data.name
    )
    
    doc = tutor_obj.model_dump()
    doc['password'] = hash_password(tutor_data.password)
    
    await db.tutors.insert_one(doc)
    
    # Create JWT token
    token = create_jwt_token(tutor_obj.id)
    
    return {
        "tutor": tutor_obj.model_dump(),
        "token": token
    }

@api_router.post("/tutors/login")
async def login_tutor(login_data: TutorLogin):
    # Find tutor
    tutor = await db.tutors.find_one({"email": login_data.email})
    if not tutor:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Verify password
    if not verify_password(login_data.password, tutor['password']):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Create token
    token = create_jwt_token(tutor['id'])
    
    # Return tutor data without password
    tutor_data = {k: v for k, v in tutor.items() if k != 'password' and k != '_id'}
    
    return {
        "tutor": tutor_data,
        "token": token
    }

@api_router.get("/tutors/me")
async def get_current_tutor_info(tutor_id: str = Depends(get_current_tutor)):
    tutor = await db.tutors.find_one({"id": tutor_id}, {"_id": 0, "password": 0})
    if not tutor:
        raise HTTPException(status_code=404, detail="Tutor not found")
    return tutor

# ============ EXAM ROUTES ============

@api_router.post("/exams", response_model=Exam)
async def create_exam(exam_data: ExamCreate, tutor_id: str = Depends(get_current_tutor)):
    # Convert QuestionCreate to Question
    questions = [Question(**q.model_dump()) for q in exam_data.questions]
    
    exam_obj = Exam(
        tutor_id=tutor_id,
        title=exam_data.title,
        description=exam_data.description,
        required_fields=exam_data.required_fields,
        questions=questions,
        settings=exam_data.settings
    )
    
    doc = exam_obj.model_dump()
    await db.exams.insert_one(doc)
    
    return exam_obj

@api_router.get("/exams", response_model=List[Exam])
async def get_tutor_exams(tutor_id: str = Depends(get_current_tutor)):
    exams = await db.exams.find({"tutor_id": tutor_id}, {"_id": 0}).to_list(1000)
    return exams

@api_router.get("/exams/{exam_id}", response_model=Exam)
async def get_exam(exam_id: str, tutor_id: str = Depends(get_current_tutor)):
    exam = await db.exams.find_one({"id": exam_id, "tutor_id": tutor_id}, {"_id": 0})
    if not exam:
        raise HTTPException(status_code=404, detail="Exam not found")
    return exam

@api_router.put("/exams/{exam_id}", response_model=Exam)
async def update_exam(exam_id: str, exam_data: ExamCreate, tutor_id: str = Depends(get_current_tutor)):
    # Check if exam exists and belongs to tutor
    existing = await db.exams.find_one({"id": exam_id, "tutor_id": tutor_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Exam not found")
    
    questions = [Question(**q.model_dump()) for q in exam_data.questions]
    
    update_data = {
        "title": exam_data.title,
        "description": exam_data.description,
        "required_fields": exam_data.required_fields,
        "questions": [q.model_dump() for q in questions],
        "settings": exam_data.settings.model_dump()
    }
    
    await db.exams.update_one({"id": exam_id}, {"$set": update_data})
    
    updated_exam = await db.exams.find_one({"id": exam_id}, {"_id": 0})
    return updated_exam

@api_router.delete("/exams/{exam_id}")
async def delete_exam(exam_id: str, tutor_id: str = Depends(get_current_tutor)):
    result = await db.exams.delete_one({"id": exam_id, "tutor_id": tutor_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Exam not found")
    return {"message": "Exam deleted successfully"}

# ============ STUDENT EXAM ROUTES (NO AUTH) ============

@api_router.get("/exams/{exam_id}/public")
async def get_public_exam(exam_id: str):
    exam = await db.exams.find_one({"id": exam_id, "is_active": True}, {"_id": 0})
    if not exam:
        raise HTTPException(status_code=404, detail="Exam not found or inactive")
    
    # Check if exam is within date range
    settings = exam.get('settings', {})
    now = datetime.now(timezone.utc).isoformat()
    
    if settings.get('start_date') and now < settings['start_date']:
        raise HTTPException(status_code=403, detail="Exam has not started yet")
    
    if settings.get('end_date') and now > settings['end_date']:
        raise HTTPException(status_code=403, detail="Exam has ended")
    
    # Remove correct answers from questions
    exam_copy = exam.copy()
    for question in exam_copy['questions']:
        question.pop('correct_answer', None)
    
    return exam_copy

@api_router.post("/exams/{exam_id}/submit")
async def submit_exam(exam_id: str, submission: ExamSubmission):
    # Get exam
    exam = await db.exams.find_one({"id": exam_id})
    if not exam:
        raise HTTPException(status_code=404, detail="Exam not found")
    
    # Calculate score
    score = 0
    max_score = 0

    for question in exam['questions']:
        # Only count auto-graded questions (those with a correct_answer present)
        q_correct = question.get('correct_answer', None)
        if q_correct is None or q_correct == '':
            # skip open-ended questions from auto-grading
            continue

        max_score += question.get('points', 0)

        # Find student's answer
        student_answer = next((a for a in submission.answers if a.question_id == question['id']), None)
        if not student_answer:
            continue

        # Handle list correct answers (multiple_select)
        if isinstance(q_correct, list):
            # student's answer may be a JSON string representing a list
            student_val = student_answer.answer
            try:
                parsed = json.loads(student_val) if isinstance(student_val, str) else student_val
            except Exception:
                parsed = None
            if isinstance(parsed, list):
                # compare as sets (after trimming)
                a = sorted([str(x).strip() for x in q_correct])
                b = sorted([str(x).strip() for x in parsed])
                if a == b:
                    score += question.get('points', 0)
        else:
            try:
                # numeric questions: compare numerically
                if question.get('type') == 'numeric':
                    try:
                        s = float(student_answer.answer)
                        c = float(q_correct)
                        if s == c:
                            score += question.get('points', 0)
                    except Exception:
                        pass
                else:
                    if str(student_answer.answer).strip().lower() == str(q_correct).strip().lower():
                        score += question.get('points', 0)
            except Exception:
                pass
    
    percentage = (score / max_score * 100) if max_score > 0 else 0
    
    # Check if flagged (too many violations)
    flagged = len(submission.violations) >= exam['settings']['max_violations']
    
    # Create attempt
    attempt = ExamAttempt(
        exam_id=exam_id,
        student_data=submission.student_data,
        answers=submission.answers,
        violations=submission.violations,
        score=score,
        max_score=max_score,
        percentage=percentage,
        flagged=flagged,
        browser_info=submission.browser_info,
        ip_address=submission.ip_address
    )
    
    doc = attempt.model_dump()
    await db.exam_attempts.insert_one(doc)
    
    return {
        "attempt_id": attempt.id,
        "score": score,
        "max_score": max_score,
        "percentage": percentage,
        "flagged": flagged,
        "violations_count": len(submission.violations)
    }

@api_router.post("/exams/{exam_id}/violations")
async def report_violation(exam_id: str, report: ViolationReport):
    # Just log it for now, could be used for real-time monitoring
    await db.violation_logs.insert_one({
        "exam_id": exam_id,
        "violation": report.violation.model_dump(),
        "logged_at": datetime.now(timezone.utc).isoformat()
    })
    return {"message": "Violation logged"}

# ============ RESULTS ROUTES ============

@api_router.get("/exams/{exam_id}/attempts")
async def get_exam_attempts(exam_id: str, tutor_id: str = Depends(get_current_tutor)):
    # Verify exam belongs to tutor
    exam = await db.exams.find_one({"id": exam_id, "tutor_id": tutor_id})
    if not exam:
        raise HTTPException(status_code=404, detail="Exam not found")
    
    attempts = await db.exam_attempts.find({"exam_id": exam_id}, {"_id": 0}).to_list(1000)
    return attempts

@api_router.get("/exams/{exam_id}/analytics")
async def get_exam_analytics(exam_id: str, tutor_id: str = Depends(get_current_tutor)):
    # Verify exam belongs to tutor
    exam = await db.exams.find_one({"id": exam_id, "tutor_id": tutor_id})
    if not exam:
        raise HTTPException(status_code=404, detail="Exam not found")
    
    attempts = await db.exam_attempts.find({"exam_id": exam_id}, {"_id": 0}).to_list(1000)
    
    if not attempts:
        return {
            "total_attempts": 0,
            "average_score": 0,
            "flagged_count": 0,
            "completion_rate": 0
        }
    
    total = len(attempts)
    avg_score = sum(a['percentage'] for a in attempts) / total
    flagged = sum(1 for a in attempts if a['flagged'])
    
    return {
        "total_attempts": total,
        "average_score": round(avg_score, 2),
        "flagged_count": flagged,
        "completion_rate": 100.0,
        "highest_score": max(a['percentage'] for a in attempts),
        "lowest_score": min(a['percentage'] for a in attempts)
    }

@api_router.get("/exams/{exam_id}/export")
async def export_exam_results(exam_id: str, tutor_id: str = Depends(get_current_tutor)):
    # Verify exam belongs to tutor
    exam = await db.exams.find_one({"id": exam_id, "tutor_id": tutor_id})
    if not exam:
        raise HTTPException(status_code=404, detail="Exam not found")
    
    attempts = await db.exam_attempts.find({"exam_id": exam_id}, {"_id": 0}).to_list(1000)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Exam Results"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Define headers based on required fields
    required_fields = exam.get('required_fields', [])
    headers = []
    for field in required_fields:
        headers.append(field.replace('_', ' ').title())
    
    headers.extend(['Score', 'Max Score', 'Percentage', 'Violations', 'Flagged', 'Submitted At'])
    
    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write data
    for row_idx, attempt in enumerate(attempts, start=2):
        col = 1
        # Student data
        for field in required_fields:
            ws.cell(row=row_idx, column=col, value=attempt['student_data'].get(field, ''))
            col += 1
        
        # Scores
        ws.cell(row=row_idx, column=col, value=attempt['score'])
        col += 1
        ws.cell(row=row_idx, column=col, value=attempt['max_score'])
        col += 1
        ws.cell(row=row_idx, column=col, value=f"{attempt['percentage']:.2f}%")
        col += 1
        ws.cell(row=row_idx, column=col, value=len(attempt['violations']))
        col += 1
        ws.cell(row=row_idx, column=col, value='Yes' if attempt['flagged'] else 'No')
        col += 1
        ws.cell(row=row_idx, column=col, value=attempt['submitted_at'])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Return as streaming response
    filename = f"{exam['title'].replace(' ', '_')}_results.xlsx"
    headers = {
        'Content-Disposition': f'attachment; filename="{filename}"'
    }
    
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers=headers
    )

@api_router.get("/")
async def root():
    return {"message": "ExamShield API is running"}

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


